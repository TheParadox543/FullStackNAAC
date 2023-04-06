import logging
from json import dumps
from os import system as ossystem
from sys import exit as sysexit
from typing import Optional, TypeVar

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter as get_col_let
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Using the logs.
logger_monitor = logging.getLogger(__name__)
logger_monitor.setLevel(logging.ERROR)
handler = logging.FileHandler("drive_reader_logs.log")
handler.setFormatter(logging.Formatter("%(asctime)s:%(levelname)s:%(name)s: %(message)s"))
logger_monitor.addHandler(handler)


# * Declare a few types to help with understanding.
Category = TypeVar("Category", bound=str)
Classification = TypeVar("Classification", bound=str)
Code = TypeVar("Code", bound=str)
Name = TypeVar("Name", bound=str)
Year = TypeVar("Year", bound=str)


def sort_dictionary(unsorted_dict: dict[str, ], reverse=False):
    """A util function to sort the keys of a dictionary.

    Parameters
    ---------
    - unsorted_dict`dict[str, Any]`: The dictionary that needs to be sorted.
    - reverse`bool`: Whether the keys need to be sorted in reverse order.

    Returns
    ------
    - sorted_dict`dict[str, Any]`: The dictionary with keys sorted.
    """
    order_list: list[str] = sorted(unsorted_dict.keys(), reverse=reverse)
    sorted_dictionary = dict({i: unsorted_dict[i] for i in order_list})
    return sorted_dictionary


class ExcelWorker():
    """The class that will handle interaction with the excel workbooks."""

    def __init__(self) -> None:
        self.read_classification_exl()

    def read_classification_exl(self):
        """Read the classification categories."""
        # Load the workbook, if not exit the program.
        try:
            self.doc_wb:Workbook = load_workbook("data/doc_classification.xlsx")
            logger_monitor.debug("Classification file found.")
        except FileNotFoundError:
            print("Classification file not found.")
            logger_monitor.warning("Classification file not found.")
            sysexit()

        # * The following are the data to be extracted from the excel sheet.
        # Category - The name of major types of requirements.
        # Classification - The numeric code.
        # Code - A 4 letter (sometimes 3) that represents what it is.
        # Name - The full name of the previous mentioned.

        # "RPIF": [full name, category, [1.2, 2,3]]
        self.code_list: dict[Code, tuple[Name, Category, list[Classification]]] = {}
        # "2.3.4": "Research"
        self.classification_list: dict[Classification, Category] = {}
        ws: Worksheet

        # Loop through all sheets and categorize each code.
        for ws in self.doc_wb:
            if ws.title == "NAAC Quantitative":

                for row in ws.iter_rows(min_col=1, max_col=4,
                                                min_row=3, values_only=True):
                    category: Category = row[0] or category
                    classification: Classification = row[1] or classification
                    code: Optional[Code] = row[2]
                    name: Optional[Name] = row[3]

                    if code and name:
                        if code not in self.code_list:
                            self.code_list[code] = [name, category,
                                                    [classification]]
                        else:
                            self.code_list[code][2].append(classification)
                    if classification not in self.classification_list:
                        self.classification_list[classification] = category

            #Any other sheet, check if any code is missing and then add it.
            else:
                for row in ws.iter_rows(min_col=2, max_col=3):
                    code, name = row[0].value, row[1].value
                    if code and name and code not in self.code_list:
                        self.code_list[code] = [name, ws.title, ["Unknown"]]

        #Write the generated data to files for evaluation.
        with open("data/code_list.json", "w") as file:
            code_obj = dumps(self.code_list, indent=4)
            file.write(code_obj)
        with open("data/classification_list.json", "w") as file:
            class_obj = dumps(self.classification_list, indent=4)
            file.write(class_obj)
        logger_monitor.debug(self.code_list)
        logger_monitor.debug(self.classification_list)

    def write_data_to_excel(self,
            drive_data: dict[Category, dict[Year, dict[Code, int]]],
            exempted: list[tuple[Name, str]]):
        """Write data from the drive to the excel sheet.

        Parameters
        ----------
        - drive_data: The raw data of the different files that satisfy
        the necessary conditions in all folders.
        - exempted: The files that are exempted from classification
        due to any reason.
        """
        workbook = Workbook()
        workbook.active.title = "exempted"

        # Loop through all the categories to create new sheets.
        for category in drive_data:
            worksheet: Worksheet = workbook.create_sheet(category, -1)
            category_data = drive_data[category]
            worksheet.append(["YEAR", "CLASSIFICATION", "COUNT"])
            # Format headers in each sheet.
            for i in range(1, 4):
                worksheet[f"{get_col_let(i)}1"].alignment = Alignment(horizontal="center")
                worksheet[f"{get_col_let(i)}1"].font = Font(bold=True, size=12)

            # Append data to the sheet.
            start, stop, width = 2, 2, 16
            for year, year_data in category_data.items():
                worksheet[f"A{start}"] = year
                for code, val in year_data.items():
                    name = self.code_list[code][0]
                    worksheet[f"B{stop}"] = name
                    worksheet[f"C{stop}"] = val
                    width = max(width, len(name))
                    stop += 1

                # Merge the cells of same years, and center the alignment.
                worksheet.merge_cells(f"A{start}:A{stop-1}")
                worksheet[f"A{start}"].alignment = Alignment(horizontal="center",
                                                             vertical="center")
                start = stop
            # Fix width to readable length.
            worksheet.column_dimensions["B"].width = width
            worksheet.column_dimensions["A"].width = 13

        # Handle exempted data.
        worksheet = workbook["exempted"]
        worksheet.append(["File Name", "Folder name"])
        for i in range(1, 3):
            worksheet[f"{get_col_let(i)}1"].alignment = Alignment(horizontal="center")
            worksheet[f"{get_col_let(i)}1"].font = Font(bold=True, size=12)
        width1, width2 = 13, 13
        # Append data to the sheet.
        for value in exempted:
            worksheet.append(value)
            width1, width2 = max(width1, len(value[0])), max(width2, len(value[1]))
        # Fix width to readable length
        worksheet.column_dimensions["A"].width = width1
        worksheet.column_dimensions["B"].width = width2

        # Save the workbook, close any instance if saving fails.
        while True:
            try:
                workbook.save("data/categorized.xlsx")
            except PermissionError:
                try:
                    ossystem("taskkill/im EXCEL.EXE categorized.xlsx")
                except:
                    pass
            else:
                break
        # * Open the workbook to see the result.
        ossystem("start EXCEL.EXE data/categorized.xlsx")

    def write_naac_data_to_excel(self,
            drive_data: dict[Category, dict[Year, dict[Code, int]]]):
        """Write data to excel sheet in naac required format.

        Parameters
        ----------
        - drive_data: The raw data of the different files that satisfy
        the necessary conditions in all folders.
        """
        spec_data = {} # Alias for classification data.

        # Take the data for 22-23 and put in naac excel.
        for category_data in drive_data.values():
            for year, year_data in category_data.items():
                if year == "2022-2023":
                    for code, value in year_data.items():
                        if self.code_list.get(code):
                            for spec in self.code_list[code][2]:
                                spec_data.update({
                                    spec: spec_data.get(spec, 0) + value
                                })
        spec_data = sort_dictionary(spec_data)

        naac_wb = Workbook()
        naac_ws: Worksheet = naac_wb.active
        naac_ws.title = "2022-2023"
        start, width = 1, 13
        logger_monitor.debug(spec_data)

        # * Entering the data that is there.
        # for spec in spec_data:
        #     number = int(spec[0])
        #     if old_number != number:
        #         naac_ws.merge_cells(f"A{start}:A{stop}")
        #         start = stop + 1
        #         naac_ws[f"A{start}"].alignment = Alignment(horizontal="center",
        #                                                    vertical="center")
        #         old_number = number
        #         # stop += 1
        #         naac_ws[f"A{start}"] = self.categories[number]
        #         width = max(width, len(self.categories[number])*1.3)
        #     stop += 1
        #     naac_ws[f"B{stop}"] = spec
        #     naac_ws[f"C{stop}"] = spec_data[spec]
        # naac_ws.column_dimensions["A"].width = width

        # * Entering all specification codes.
        naac_ws.append(["Classification", "Code", "Count"])
        start, word = 1, "Classification"
        for num, (classification, category) in \
                enumerate(self.classification_list.items(), 2):
            naac_ws.append({
                    "B": classification,
                    "C": spec_data.get(classification, 0)
            })
            width = max(width, len(category)*1.2)
            # When word changes, merge cells.
            if word != category:
                naac_ws.merge_cells(f"A{start}:A{num-1}")
                naac_ws[f"A{start}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                naac_ws[f"A{start}"] = word
                start, word = num, category
        else:
            # Merge remaining categories.
            naac_ws.merge_cells(f"A{start}:A{num}")
            naac_ws[f"A{start}"].alignment = Alignment(horizontal="center",
                                                        vertical="center")
            naac_ws[f"A{start}"] = category

            # Fix alignment and font for headers.
            for i in range(1, 4):
                naac_ws[f"{get_col_let(i)}1"].alignment = Alignment(
                    horizontal="center", vertical="center")
                naac_ws[f"{get_col_let(i)}1"].font = Font(bold=True, size=12)
            # Improve readability of category column.
            naac_ws.column_dimensions["A"].width = width

        # Save the workbook, close any instances if saving fails.
        while True:
            try:
                naac_wb.save("data/naac.xlsx")
                break
            except PermissionError:
                logger_monitor("Failed to save naac.xlsx")
                ossystem("taskkill /im EXCEL.EXE naac.xlsx")
        ossystem("start EXCEL.EXE data/naac.xlsx")
